# locust --autostart --autoquit 2 -f run.py
# python -m locust --autostart --autoquit 2 -f "D:\pytest\测试\run.py"
import base64
from datetime import datetime
from locust import task, HttpUser, events, LoadTestShape
import logging
import time
import json
import random
import os
import yaml
from gevent.event import Event
import gevent
import sqlite3
import pandas as pd
import re

log_print_rate = 1


class Utils:
    @staticmethod
    def sql_execute(file_name, sql, param):
        conn = sqlite3.connect(f'{file_name}.db')
        cursor = conn.cursor()
        if param is not None:
            cursor.execute(sql, param)
        else:
            cursor.execute(sql)
        conn.commit()
        conn.close()

    @staticmethod
    def init_cache_file(file_name):
        create_table_sql = '''CREATE TABLE IF NOT EXISTS results
                              (
                                  model_name
                                  TEXT,
                                  input_length
                                  INTEGER,
                                  output_length
                                  INTEGER,
                                  user_count
                                  INTEGER,
                                  question
                                  TEXT,
                                  question_len
                                  INTEGER,
                                  answer
                                  TEXT,
                                  answer_len
                                  INTEGER,
                                  prompt_tokens
                                  INTEGER,
                                  completion_tokens
                                  INTEGER,
                                  ttft
                                  REAL,
                                  tps_decode
                                  REAL,
                                  tpot_decode
                                  REAL,
                                  cps_decode
                                  REAL,
                                  latency
                                  REAL,
                                  tps_all
                                  REAL,
                                  tpot_all
                                  REAL,
                                  cps_all
                                  REAL,
                                  prefill_count
                                  INTEGER,
                                  decode_count
                                  INTEGER,
                                  question_num
                                  INTEGER,
                                  running_request_count
                                  INTEGER
                              );'''
        Utils.sql_execute(file_name, create_table_sql, None)

    # 定义非法字符的模式（Excel不允许的控制字符）
    ILLEGAL_CHAR_PATTERN = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')

    @staticmethod
    def has_illegal_chars(val):
        """检查字符串是否包含非法字符"""
        if isinstance(val, str):
            return bool(Utils.ILLEGAL_CHAR_PATTERN.search(val))
        return False  # 非字符串类型认为合法

    @staticmethod
    def remove_rows_with_illegal_chars(df):
        """
        删除 DataFrame 中包含非法字符的行
        """
        # 创建一个布尔掩码，标记所有“合法”的行
        mask = pd.Series([True] * len(df), index=df.index)

        for col in df.columns:
            if df[col].dtype == 'object':  # 只检查字符串列
                col_has_illegal = df[col].apply(Utils.has_illegal_chars)
                mask &= ~col_has_illegal  # 去掉包含非法字符的行

        cleaned_df = df[mask].copy()
        return cleaned_df

    @staticmethod
    def db2excel(file_name):
        conn = sqlite3.connect(f'{file_name}.db')
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        with pd.ExcelWriter(f'{file_name}.xlsx', engine='openpyxl') as writer:
            for table_name in tables:
                table = table_name[0]
                df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
                if df.empty:
                    df.to_excel(writer, sheet_name=table, index=False)
                    continue
                df_cleaned = Utils.remove_rows_with_illegal_chars(df)
                df_cleaned.to_excel(writer, sheet_name=table, index=False)
        conn.close()

    @staticmethod
    def deep_get(data, path, default=None):
        keys = path.split('.')
        for key in keys:
            if isinstance(data, dict) and key in data:
                data = data[key]
            elif isinstance(data, list) and key.isdigit() and int(key) < len(data):
                data = data[int(key)]
            else:
                return default
        return data

    @staticmethod
    def count_data(file_name):
        global result_json
        result_json = {'model_name': model_name}

        df = pd.read_excel(f"{file_name}.xlsx", sheet_name="results")
        # df = df[df['ttft'] > 0]  # 删除无效数据
        # users_count_list = df['user_count'].dropna().drop_duplicates().tolist()
        users_count_list = fixed_user_count
        input_length_list = df['input_length'].dropna().drop_duplicates().tolist()

        data = []
        for _input_length in input_length_list:
            result_json[f"input_{_input_length}"] = []
            for _user_count in users_count_list:
                df_round = df[(df['input_length'] == _input_length) & (df['user_count'] == _user_count)]
                total_count = max(len(df_round), 1)
                fail_count = len(df_round[(df_round['ttft'] <= 0) | (df_round['completion_tokens'] <= 0)])

                df_round = df_round[(df_round['ttft'] > 0) & (df_round['completion_tokens'] > 0)]  # 删除无效数据

                filtered_df = df_round[df_round['running_request_count'] == _user_count]  # 其余数据

                ttft_p20 = df_round[df_round['prefill_count'] <= max(int(_user_count) * 0.2, 1)]['ttft']
                ttft_p50 = df_round[df_round['prefill_count'] <= max(int(_user_count) * 0.5, 1)]['ttft']
                ttft_p75 = df_round[df_round['prefill_count'] <= max(int(_user_count) * 0.75, 1)]['ttft']
                ttft_avg = df_round[df_round['prefill_count'] <= max(int(_user_count), 1)]['ttft']

                def safe_mean(series, decimal=4):
                    valid = series.dropna()
                    return round(valid.mean(), decimal) if not valid.empty else 0.0

                p20_real_ttft = safe_mean(ttft_p20)
                p50_real_ttft = safe_mean(ttft_p50)
                p75_real_ttft = safe_mean(ttft_p75)
                avg_ttft = safe_mean(ttft_avg)

                result_data = {
                    "平均输入字符数": _input_length,
                    "并发/指标": _user_count,
                    "p20_real_ttft": p20_real_ttft,
                    "p50_real_ttft": p50_real_ttft,
                    "p75_real_ttft": p75_real_ttft,
                    "avg_ttft": avg_ttft,
                    "Single_TPS": round(filtered_df['tps_decode'].mean(), 4),
                    "Single_CPS": round(filtered_df['cps_decode'].mean(), 4),
                    "Single_TPS_A": round(filtered_df['tps_all'].mean(), 4),
                    "Single_CPS_A": round(filtered_df['cps_all'].mean(), 4),
                    "TPS": round(filtered_df['tps_decode'].mean() * _user_count, 4),
                    "CPS": round(filtered_df['cps_decode'].mean() * _user_count, 4),
                    "TPS_A": round(filtered_df['tps_all'].mean() * _user_count, 4),
                    "CPS_A": round(filtered_df['cps_all'].mean() * _user_count, 4),
                    "TPOT": round(filtered_df['tpot_decode'].mean(), 4),
                    "平均输入token数": round(filtered_df['prompt_tokens'].mean(), 0),
                    "平均输出token数": round(filtered_df['completion_tokens'].mean(), 0),
                    "总请求数": total_count,
                    "失败请求数": fail_count,
                    "请求成功率": round((total_count - fail_count) / total_count, 4)
                }
                data.append(result_data)
                result_json[f"input_{_input_length}"].append({
                    f"users_{_user_count}": result_data
                })
        df = pd.DataFrame(data)
        with pd.ExcelWriter(f"result{str(fixed_user_count[0])}.xlsx", engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)

            from openpyxl.styles import Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            # 获取 workbook 和 worksheet
            wb = writer.book
            ws = wb["Sheet1"]

            # 设置全表居中
            alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 遍历所有单元格进行格式设置
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = alignment
                    cell.border = border

            # 合并第一列中相同值的单元格（仅合并内容相同的连续行）
            prev_value = None
            start_row = None
            merge_col = 1  # A列

            for row_idx in range(2, ws.max_row + 2):  # 从第2行开始（数据行）
                current_value = ws.cell(row=row_idx, column=merge_col).value

                if current_value != prev_value:
                    # 合并上一组
                    if start_row and start_row != row_idx - 1:
                        ws.merge_cells(start_row=start_row, start_column=merge_col,
                                       end_row=row_idx - 1, end_column=merge_col)
                        # 居中合并后的单元格
                        ws.cell(start_row, merge_col).alignment = alignment
                    # 新组开始
                    start_row = row_idx
                prev_value = current_value

            # 处理最后一组
            if start_row and start_row < ws.max_row + 1:
                ws.merge_cells(start_row=start_row, start_column=merge_col,
                               end_row=ws.max_row + 1, end_column=merge_col)
                ws.cell(start_row, merge_col).alignment = alignment

            def apply_thick_border(ws, start_row, end_row, start_col, end_col):
                thin = Side(style='thin')
                medium = Side(style='medium')
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = ws.cell(row=row, column=col)
                        border = Border(
                            left=thin if col > start_col else medium,  # 非最左列：细线；最左列：粗线
                            right=thin if col < end_col else medium,  # 非最右列：细线；最右列：粗线
                            top=thin if row > start_row else medium,  # 非最上行：细线；最上行：粗线
                            bottom=thin if row < end_row else medium  # 非最下行：细线；最下行：粗线
                        )
                        cell.border = border

            apply_thick_border(ws, start_row=1, end_row=1, start_col=1, end_col=20)
            for i in range(len(input_length)):
                apply_thick_border(ws, start_row=len(users_count_list) * i + 2,
                                   end_row=len(users_count_list) * (i + 1) + 1,
                                   start_col=1, end_col=20)
            # 调整列宽（美观）
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 16

    @staticmethod
    def test_set_creat(arg_input_length, arg_output_length):
        questions = []

        demo = '一、引言：提出主题，激发兴趣开篇通过引用名言、列举数据、讲述小故事或提出问题等方式，引出文章核心主题。简要说明该话题的现实意义或社会关注度，明确写作目的。例如：“在数字化浪潮席卷全球的今天，人工智能正深刻改变着我们的生活方式。”随后，用一句话概括全文中心论点或主要观点，为下文铺垫。二、背景介绍：梳理发展脉络介绍主题的历史渊源、发展过程或基本概念，帮助读者建立认知基础。可结合时间线、关键事件或政策演变，展现其演进逻辑。例如，阐述某项技术的起源、某社会现象的形成原因等，增强文章的深度与广度。三、现状分析：呈现当前情况描述该主题在当前社会中的实际表现，包括主流应用、典型模式、普及程度等。可通过具体案例、统计数据或调查结果进行支撑，使论述更具说服力。同时指出存在的普遍问题或公众关注的焦点。四、主体论述：分点深入分析将核心内容分为2-3个子主题，每部分独立成段，采用“观点+论据+例证+小结”的结构展开。可运用对比、因果、分类、举例等论证方法，确保逻辑严密、层次分明。例如，从技术、经济、伦理等不同维度进行剖析。五、问题与挑战：客观反思指出当前发展中的瓶颈、风险或争议点，如技术局限、安全隐患、公平性问题等。分析其成因，体现批判性思维和全面视角。六、对策与建议：提出解决方案针对前述问题，提出切实可行的改进措施或发展建议。可从政策制定、技术创新、公众参与、教育引导等方面入手，突出建设性和前瞻性。七、结论与展望：总结升华总结全文核心观点，呼应开头。对未来发展趋势进行展望，呼吁行动或引发思考，增强文章的感染力与影响力。八、参考文献（可选）列出主要参考资料，提升学术规范性与可信度。'

        fields = ['历史', '文化', '旅游', '娱乐', '美食', '情感', '摄影', '生活', '时事', '健康', '时尚', '育儿',
                  '科普', '教育', '星座', '财经', '体育', '科技', '社会', '动漫']

        styles = ['通俗易懂', '专业严谨', '生动形象', '幽默风趣', '庄重典雅', '简洁明快', '抒情细腻', '议论深刻',
                  '叙事性强', '客观中立']

        for field in fields:
            for style in styles:
                if output_length != 0:
                    question = f"你是一位创作领域的专家。请撰写一篇{field}相关的文章，面向普通群众读者。要求语言{style}。文章总字数大约{arg_output_length}个汉字。"
                else:
                    question = f"你是一位创作领域的专家。请撰写一篇{field}相关的文章，面向普通群众读者。要求语言{style}。文章不限制总字数。"
                question = f"{question}；可以参考以下大纲生成，但也不必完全按照大纲生成。参考大纲：{demo * 50}"
                questions.append(question[:arg_input_length])
        return questions


config_path = os.getenv("TOOL_CONFIG")
if config_path is None:
    config_path = "config.yml"

pvc_path = "/".join(config_path.split('/')[:-1])

with open(config_path, 'r') as file:
    config = yaml.safe_load(file)

# 模型名称
model_name = config.get("model_name", "HaiRuo-7B-General-V1.0.0.0")
# 输入长度 任意指定 一般为128（短文本）1024（普通文本）4096（长文本）30000（超长文本）
input_length = [i.strip() for i in config.get("input_length", "128").split(",")]
# 输出长度限制 非硬限制 仅在prompt中做软限定 为0时不做限制
output_length = config.get("output_length", "0")
# http请求地址
hosts = config.get("host", "http://localhost:80")
# http请求路径
base_url = config.get("base_url", "/v1/chat/completions")
# 是否流失请求
is_stream = config.get("is_stream", True)
# 是否关注token
is_tokens_ness = config.get("is_tokens_ness", True)
# 请求header 可自定义传入
headers = config.get("headers", {
    "content-type": "application/json",
    "authorization": "7c3eafb5-2d6e-100d-ab0f-7b2c1cdafb3c"
})
# 请求body 可自定义传入 自定义传入时需手动传入模型名 并添加$question占位符以替换请求问题
body_template = config.get("body", {
    "model": model_name,
    "stream": True,
    "messages": [
        {
            "role": "user",
            "content": "$question"
        }
    ],
    "stream_options": {
        "include_usage": True,
        "continuous_usage_stats": True
    }
})
# 请求结果控制配置参数
config_control_param = json.loads("{}")
# 响应过程中是否有连接检查chunk
config_control_param.update({"connect_check": config.get("connect_check", True)})
# 连接检查chunk包含的字符串
config_control_param.update({"connect_check_string": config.get("connect_check_string", "ping")})
# 响应结果是否有终止标识
config_control_param.update({"done_flag": config.get("done_flag", True)})
# 响应终止chunk包含的字符串
config_control_param.update({"done_flag_string": config.get("done_flag_string", "[DONE]")})
# token统计方式 直接在最后请求的json中回显用json 自动统计用auto
config_control_param.update({"token_count_method": config.get("token_count_method", "json")})
# 流式请求最后一个token结束标识
config_control_param.update({"finish_flag": config.get("finish_flag", "stop")})
# 流式请求最后一个token结束标识位置
config_control_param.update({"finish_flag_string": config.get("finish_flag_string", "choices.0.finish_reason")})
# 输入token使用json方式统计，jsonpath位置
config_control_param.update({"prompt_tokens_string": config.get("prompt_tokens_string", "usage.prompt_tokens")})
# 输出token使用json方式统计，jsonpath位置
config_control_param.update(
    {"completion_tokens_string": config.get("completion_tokens_string", "usage.completion_tokens")})
# 流式请求内容content的jsonpath位置
config_control_param.update({"content_string": config.get("content_string", "choices.0.delta.content")})
# 文件导出位置及名称 采用模型-时间戳命名
export_file_name = f"{model_name}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
# 结果json存储dict
result_json = {}

# 每轮次实际测试时长
single_running_time = config.get("single_running_time", 60)
# 指定TTFT下最大并发寻优的TTFT上限
max_first_char_time = config.get("max_first_char_time", 3)
# 指定TTFT下最大并发寻优的TTFT下限
min_first_char_time = config.get("min_first_char_time", 2.9)
# 固定测试并发数，逗号分隔
fixed_user_count = [int(i.strip()) for i in config.get("fixed_user_count", "1,2").split(",")]
# 动态测试并发寻优轮次
auto_user_count_times = config.get("auto_user_count_times", 0)
# 最大有效测试时长 自动计算
max_running_time = single_running_time * (len(fixed_user_count) + auto_user_count_times) * len(input_length)
# 脚本最大运行时长 防止意外无限运行
script_max_running_time = config.get("script_max_running_time", 21600)

ttft_real_percentage = config.get("ttft_real_percentage", 0.2)

stop_failed_request_num = config.get("stop_failed_request_num", 100)  # 异常请求达到阈值后立刻停止继续测试

spawn_rate = config.get("spawn_rate", 50)  # 线程任务启动速率，默认50，根据测试机器性能配置，推荐100以内，一般无需修改

# # # # # # # # # # # # # # # 内置变量 # # # # # # # # # # # # # # #

current_user_count_index = 0  # 当前测试并发索引
current_user_count = fixed_user_count[current_user_count_index]  # 当前执行的并发数，默认值为固定测试并发首项
get_ttft = 0  # 当前并发下首Token时延数据，默认0秒，由上轮次测试结束后计算
test_started_flag = False  # 测试开始标识
request_started_flag = False  # 请求开始发送标识
collection_stop_flag = False  # 集合停止标识
stop_task_num = 0  # 已集合停止的任务个数

# 程序等待时间
waiting_time = 0

# 测试集列表 根据输入长度自动生成
question_list = Utils.test_set_creat(int(input_length[0]), output_length)

# 当前测试输入长度索引
current_input_length_index = 0

prefill_requests_num = 0  # prefill阶段请求数
decoding_requests_num = 0  # decoded阶段请求数

# 当前轮次请求数记录
sent_requests_num = 0  # 已发送请求数
completed_requests_num = 0  # 已结束请求数

running_task_num = 0  # 运行中的线程数

# 已发送问题总数
question_num = 0

failed_request_num = 0

start_event = Event()


def http_client(self, method, url, body, header, stream, control_param):
    connect_check = control_param.get('connect_check')
    connect_check_string = control_param.get('connect_check_string')
    done_flag = control_param.get('done_flag')
    done_flag_string = control_param.get('done_flag_string')
    token_count_method = control_param.get('token_count_method')  # token统计方式：count/json
    finish_flag = control_param.get('finish_flag')
    finish_flag_string = control_param.get('finish_flag_string')
    completion_tokens_string = control_param.get('completion_tokens_string')
    prompt_tokens_string = control_param.get('prompt_tokens_string')
    content_string = control_param.get('content_string')

    total_answer = ''
    total_time = time.time()
    first_token_time = 0
    completion_tokens = 0
    prompt_tokens = 0

    global prefill_requests_num, decoding_requests_num, sent_requests_num, completed_requests_num, failed_request_num
    start_time = time.time()
    prefill_requests_num_current = 0
    decoding_requests_num_current = 0
    init_prefill_requests_num = prefill_requests_num
    init_decoding_requests_num = decoding_requests_num
    try:
        prefill_requests_num += 1
        sent_requests_num += 1
        if stream:
            if method == 'POST':
                # 本地图片路径
                image_path = "image1.png"

                # OCR prompt，可自定义
                prompt_text = "<image>\nFree OCR."

                # 读取图片并转换为 base64
                with open(image_path, "rb") as f:
                    img_b64 = base64.b64encode(f.read()).decode("utf-8")
                # 构造 JSON 请求体
                payload = {
                    "model": "DeepSeek-OCR",  # 可填你模型名
                    "stream": True,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt_text},
                                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                            ]
                        }
                    ],
                    "response_format": {"type": "text"},
                }
                start_time = time.time()
                headers1 = {"Content-Type": "application/json"}
                with self.client.post(url, headers=headers1, data=json.dumps(payload), stream=True,
                                      catch_response=True) as response:
                    if response.status_code != 200:
                        failed_request_num += 1
                    prefill_requests_num_current = prefill_requests_num
                    decoding_requests_num_current = decoding_requests_num
                    for chunk in response.iter_lines():
                        if connect_check and connect_check_string in str(chunk):
                            continue
                        if done_flag and done_flag_string in str(chunk):
                            break
                        decoded_chunk = chunk.decode('utf-8', errors='ignore')
                        print(decoded_chunk)
                        if str(decoded_chunk).startswith("data: "):
                            data = json.loads(decoded_chunk.split('data: ')[1])
                            if first_token_time == 0:
                                prefill_requests_num -= 1
                                decoding_requests_num += 1
                                first_token_time = time.time()
                            if token_count_method == 'count':
                                completion_tokens += 1
                            else:
                                if finish_flag == Utils.deep_get(data, finish_flag_string) or len(
                                        Utils.deep_get(data, "choices")) == 0:
                                    completion_tokens = Utils.deep_get(data, completion_tokens_string)
                                    prompt_tokens = Utils.deep_get(data, prompt_tokens_string)
                            if len(Utils.deep_get(data, "choices")) != 0:
                                total_answer += str(Utils.deep_get(data, content_string))
                    total_time = time.time()
                    decoding_requests_num -= 1
                    completed_requests_num += 1
        else:
            # 本地图片路径
            # 定义目标目录
            target_dir = 'image.png'

            # 检查目录是否存在
            # if not os.path.exists(target_dir):
            #     print(f"错误：目录 '{target_dir}' 不存在")
            #     exit()
            #
            #     # 获取目录下所有png文件
            # png_files = [f for f in os.listdir(target_dir) if f.endswith('.png')]
            #
            # # 检查是否有png文件
            # if not png_files:
            #     print(f"错误：目录 '{target_dir}' 中没有找到png文件")
            #     exit()
            #
            #     # 随机选择一个文件
            # random_file = random.choice(png_files)
            #
            # # 构造完整路径
            # random_path = os.path.join(target_dir, random_file)

            # OCR prompt，可自定义
            prompt_text = "<image>\nFree OCR."

            # 读取图片并转换为 base64
            with open(target_dir, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode("utf-8")
            # 构造 JSON 请求体
            payload = {
                "model": "deepseek-ocr",  # 可填你模型名
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt_text},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                        ]
                    }
                ],
                "response_format": {"type": "text"}
            }
            start_time = time.time()
            headers1 = {"Content-Type": "application/json"}

            with self.client.post(url, headers=headers1, data=json.dumps(payload), stream=True,
                                  catch_response=True) as response:
                last_time = time.time()
                first_token_time = last_time
                total_time = time.time()
                data = json.loads(response.text)
                total_answer = str(Utils.deep_get(data, content_string))
                completion_tokens = 39
                prompt_tokens = 1
                prefill_requests_num_current = 1
                decoding_requests_num_current = 1
                completed_requests_num += 1
    except Exception as e:
        failed_request_num += 1
        total_time = time.time() + 1
        total_answer += str(e)
        completed_requests_num += 1
        completion_tokens = -1
        prompt_tokens = -1
        prefill_requests_num_current = prefill_requests_num = init_prefill_requests_num
        decoding_requests_num_current = decoding_requests_num = init_decoding_requests_num
    return first_token_time - start_time, total_time - start_time, total_answer, prompt_tokens, completion_tokens, prefill_requests_num_current, decoding_requests_num_current


class Task(HttpUser):
    host = hosts  # 获取self.host

    # 任务创建时运行一次
    def on_start(self):
        global running_task_num
        running_task_num += 1

    # 任务结束时运行一次
    def on_stop(self):
        global running_task_num
        running_task_num -= 1

        # 循环执行此任务

    @task(1)
    def task(self):
        global collection_stop_flag, stop_task_num, question_num, decoding_requests_num
        if collection_stop_flag:  # 集合停止，线程空转，不再继续发送请求，当本轮测试任务全部结束后，等待命令
            start_event.wait()
        else:
            user_count = self.environment.runner.user_count
            question = random.choice(question_list)
            question_num += 1
            current_question_num = question_num

            try:
                ttft, latency, answer, prompt_tokens, completion_tokens, prefill_requests_num_current, decoding_requests_num_current = \
                    http_client(self, 'POST', base_url, "", headers, is_stream, config_control_param)
                if latency == 0:
                    latency = 1
                if completion_tokens == 0:
                    completion_tokens = 1

                sql = '''
                      INSERT INTO results
                      (model_name, input_length, output_length, user_count,
                       question, answer, question_len, answer_len, prompt_tokens, completion_tokens,
                       ttft, tps_decode, tpot_decode, cps_decode, latency,
                       tps_all, tpot_all, cps_all, prefill_count, decode_count, question_num, running_request_count)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?); \
                      '''
                if is_stream:
                    tps_decode = completion_tokens / (latency - ttft)
                    tpot_decode = (latency - ttft) / completion_tokens
                    cps_decode = len(str(answer)) / (latency - ttft)
                else:
                    tps_decode = completion_tokens / ttft
                    tpot_decode = ttft / completion_tokens
                    cps_decode = len(str(answer)) / ttft
                param = (model_name, int(input_length[current_input_length_index]), output_length, user_count,
                         str(question), str(answer), len(str(question)), len(str(answer)), prompt_tokens,
                         completion_tokens, ttft, tps_decode, tpot_decode, cps_decode, latency,
                         completion_tokens / latency, latency / completion_tokens, len(str(answer)) / latency,
                         prefill_requests_num_current,
                         decoding_requests_num_current, current_question_num,
                         sent_requests_num - completed_requests_num + 1)
                Utils.sql_execute(export_file_name, sql, param)

            except Exception as e:
                logging.error(e)


# 继承LoadTestShape类，重构并发控制
class ConcurrencyController(LoadTestShape):
    # 每秒运行一次这个方法
    def tick(self):
        global script_max_running_time
        if self.get_run_time() > script_max_running_time:
            return None
        if failed_request_num >= stop_failed_request_num:
            return None
        if test_started_flag:
            global stop_task_num, current_user_count_index, current_user_count, get_ttft, fixed_user_count, \
                collection_stop_flag, question_list, \
                single_running_time, max_first_char_time, waiting_time, max_running_time, current_input_length_index, request_started_flag
            if collection_stop_flag or request_started_flag is False:
                waiting_time += 1
            # 当前程序运行时长
            run_time = self.get_run_time()
            if request_started_flag and int(run_time) % log_print_rate == 0:
                logging.info(
                    f"程序运行时长: {str(run_time).split('.')[0]} | 等待时长: {waiting_time} | 测试时长: {str(run_time - waiting_time).split('.')[0]} "
                    f"| 轮次: {str(current_user_count_index + (current_input_length_index * (len(fixed_user_count) + auto_user_count_times)))} "
                    f"| 并发数: {str(current_user_count)} "
                    f"| 已发送请求: {str(sent_requests_num)} | 已完成请求: {str(completed_requests_num)}")
            # 超过最大运行时长 立即停止测试
            if run_time >= max_running_time and sent_requests_num == completed_requests_num:
                return None
            max_running_time += waiting_time  # 当前程序运行时长减去等待时长，保证每轮次运行时长足够
            # 固定并发数测试阶段
            if current_user_count_index <= len(fixed_user_count):
                # 在本轮次运行时长内，继续按照本并发运行
                if run_time - waiting_time < single_running_time * (
                        current_input_length_index * len(fixed_user_count) + current_user_count_index):
                    if collection_stop_flag:
                        logging.info("5秒后开始本轮测试")
                        waiting_time += 5
                        gevent.sleep(5)  # 轮次间缓冲
                        start_event.set()
                    collection_stop_flag = False
                    request_started_flag = True
                    return current_user_count, spawn_rate
                # 当前轮次结束，轮空等待全部线程结束任务
                else:
                    # 更改集合停止标识，停止线程继续发送请求
                    collection_stop_flag = True
                    # 等待所有任务均停止后
                    if sent_requests_num == completed_requests_num:
                        current_user_count_index = current_user_count_index + 1
                        if current_user_count_index <= len(fixed_user_count):
                            current_user_count = fixed_user_count[current_user_count_index - 1]
                        else:
                            if current_input_length_index < len(input_length) - 1:
                                current_input_length_index += 1
                                question_list = Utils.test_set_creat(int(input_length[current_input_length_index]),
                                                                     output_length)
                                current_user_count_index = 1
                                current_user_count = fixed_user_count[current_user_count_index - 1]
                    # 同时维持当前并发或继续下一轮并发测试
                    return current_user_count, spawn_rate
            else:
                return None
        else:
            return 0, 50


@events.init_command_line_parser.add_listener
def _(parser):  # 全局自定义参数变量配置
    parser.add_argument('--model_name', type=str, default=model_name, help='')


@events.test_start.add_listener
def on_test_start(environment, **kwargs):  # 性能测试开始标志
    logging.info(f"测试环境: {environment}| 参数传入: {kwargs}")
    Utils.init_cache_file(export_file_name)
    logging.info("性能测试开始 ...")
    global test_started_flag
    test_started_flag = True


@events.quitting.add_listener
def on_quitting(environment, **kwargs):  # 性能测试结束标志
    logging.info(f"测试环境: {environment}| 参数传入: {kwargs}")
    Utils.db2excel(export_file_name)
    logging.info("性能测试结束 ...")
    Utils.count_data(export_file_name)
